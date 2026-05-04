"""
Admin routes:
  GET  /api/admin/db/export          → скачать дамп БД (PostgreSQL pg_dump / SQLite файл)
  POST /api/admin/db/import          → загрузить дамп БД
  GET  /api/admin/calendar/export    → скачать бронирования в Excel

Security:
  - All routes require valid JWT + admin role (require_admin dependency)
  - DB import validates file format (magic bytes) before writing
  - DB import is size-limited (MAX_IMPORT_SIZE_MB)
  - DB import backs up current DB before overwriting
  - Strict rate limiting: 5 req/minute for import, 20 req/minute for export
"""
import io
import logging
import os
import shutil
import subprocess
import tempfile
from datetime import date

from app.api.deps import get_db, require_admin
from app.core.config import settings
from app.core.limiter import limiter
from app.models.booking import Booking
from app.models.room import Room
from app.models.user import User
from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import Response, StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/admin", tags=["admin"])

# ── Security constants ────────────────────────────────────────────────────────
MAX_IMPORT_SIZE_MB = 500
MAX_IMPORT_SIZE_BYTES = MAX_IMPORT_SIZE_MB * 1024 * 1024

# Magic bytes for supported formats
SQLITE_MAGIC = b"SQLite format 3\x00"  # first 16 bytes of any SQLite3 file
PGDMP_MAGIC = b"PGDMP"               # first 5 bytes of pg_dump custom format


# ─── helpers ──────────────────────────────────────────────────────────────────

def _parse_db_url(url: str) -> dict:
    """Parse DATABASE_URL into connection components."""
    if url.startswith("postgresql"):
        raw = url.replace("postgresql+asyncpg://", "").replace("postgresql://", "")
        userpass, rest = raw.split("@", 1)
        user, password = userpass.split(":", 1)
        hostport, dbname = rest.split("/", 1)
        host, port = hostport.split(":", 1) if ":" in hostport else (hostport, "5432")
        return {"type": "postgresql", "user": user, "password": password,
                "host": host, "port": port, "dbname": dbname}
    elif url.startswith("sqlite"):
        path = url.split("///", 1)[1]
        return {"type": "sqlite", "path": path}
    else:
        raise ValueError(f"Unsupported DATABASE_URL scheme: {url}")


def _validate_sqlite(data: bytes) -> None:
    """Raise HTTPException if data is not a valid SQLite3 file."""
    if len(data) < 16 or data[:16] != SQLITE_MAGIC:
        raise HTTPException(
            400,
            detail={
                "code": "INVALID_FILE_FORMAT",
                "message": "Файл не является корректной базой данных SQLite3",
            },
        )


def _validate_pgdump(data: bytes) -> None:
    """Raise HTTPException if data is not a valid pg_dump custom-format file."""
    if len(data) < 5 or data[:5] != PGDMP_MAGIC:
        raise HTTPException(
            400,
            detail={
                "code": "INVALID_FILE_FORMAT",
                "message": "Файл не является корректным дампом PostgreSQL (custom format)",
            },
        )


def _check_size(data: bytes) -> None:
    """Raise HTTPException if file exceeds size limit."""
    if len(data) > MAX_IMPORT_SIZE_BYTES:
        raise HTTPException(
            413,
            detail={
                "code": "FILE_TOO_LARGE",
                "message": f"Файл превышает лимит {MAX_IMPORT_SIZE_MB} МБ",
            },
        )


# ─── DB Export ────────────────────────────────────────────────────────────────

@router.get("/db/export")
@limiter.limit("20/minute")
async def export_database(
    request: Request,
    current_user: User = Depends(require_admin),
):
    """Download a full database dump. Requires admin role."""
    logger.info("DB export requested by user %s", current_user.id)
    db_info = _parse_db_url(settings.DATABASE_URL)

    if db_info["type"] == "sqlite":
        path = db_info["path"]
        if not os.path.exists(path):
            raise HTTPException(
                404,
                detail={"code": "DB_NOT_FOUND", "message": "Файл БД не найден"},
            )
        with open(path, "rb") as f:
            data = f.read()
        return Response(
            content=data,
            media_type="application/octet-stream",
            headers={"Content-Disposition": "attachment; filename=hotel_manager.db"},
        )

    elif db_info["type"] == "postgresql":
        env = os.environ.copy()
        env["PGPASSWORD"] = db_info["password"]
        try:
            result = subprocess.run(
                [
                    "pg_dump",
                    "-h", db_info["host"],
                    "-p", db_info["port"],
                    "-U", db_info["user"],
                    "-d", db_info["dbname"],
                    "--no-password",
                    "--format=custom",
                ],
                env=env,
                capture_output=True,
                timeout=120,
            )
            if result.returncode != 0:
                logger.error("pg_dump failed: %s", result.stderr.decode())
                raise HTTPException(
                    502,
                    detail={"code": "PG_DUMP_ERROR", "message": result.stderr.decode()},
                )
            return Response(
                content=result.stdout,
                media_type="application/octet-stream",
                headers={"Content-Disposition": "attachment; filename=hotel_manager.dump"},
            )
        except FileNotFoundError:
            raise HTTPException(
                500,
                detail={"code": "PG_DUMP_MISSING", "message": "pg_dump не установлен"},
            )


# ─── DB Import ────────────────────────────────────────────────────────────────

@router.post("/db/import")
@limiter.limit("5/minute")
async def import_database(
    request: Request,
    file: UploadFile = File(...),
    current_user: User = Depends(require_admin),
):
    """
    Upload and restore a database dump. Requires admin role.

    Security checks:
      1. JWT + admin role (require_admin)
      2. File size limit (MAX_IMPORT_SIZE_MB)
      3. Magic bytes validation (SQLite3 / pg_dump custom format)
      4. Atomic replace with backup for SQLite
    """
    logger.warning(
        "DB import initiated by admin user %s (login=%s)",
        current_user.id,
        current_user.login,
    )

    # ── 1. Read & size check ─────────────────────────────────────────────────
    data = await file.read()
    _check_size(data)

    db_info = _parse_db_url(settings.DATABASE_URL)

    # ── 2. SQLite path ───────────────────────────────────────────────────────
    if db_info["type"] == "sqlite":
        _validate_sqlite(data)

        path = db_info["path"]
        backup_path = path + ".bak"

        # Backup current DB atomically before overwriting
        if os.path.exists(path):
            shutil.copy2(path, backup_path)
            logger.info("SQLite backup created at %s", backup_path)

        try:
            # Write to temp file first, then atomically replace
            dir_ = os.path.dirname(path) or "."
            with tempfile.NamedTemporaryFile(dir=dir_, delete=False, suffix=".tmp") as tmp:
                tmp.write(data)
                tmp_path = tmp.name
            os.replace(tmp_path, path)  # atomic on POSIX, best-effort on Windows
            logger.info("SQLite DB replaced successfully by user %s", current_user.id)
        except Exception as exc:
            # Restore backup on failure
            if os.path.exists(backup_path):
                shutil.copy2(backup_path, path)
                logger.error("SQLite import failed, backup restored: %s", exc)
            raise HTTPException(
                500,
                detail={"code": "IMPORT_ERROR", "message": f"Ошибка записи файла: {exc}"},
            )

        return {
            "ok": True,
            "message": (
                "База данных успешно загружена. "
                "Резервная копия сохранена рядом (.bak). "
                "Перезапустите сервер для применения изменений."
            ),
        }

    # ── 3. PostgreSQL path ───────────────────────────────────────────────────
    elif db_info["type"] == "postgresql":
        _validate_pgdump(data)

        env = os.environ.copy()
        env["PGPASSWORD"] = db_info["password"]

        with tempfile.NamedTemporaryFile(delete=False, suffix=".dump") as tmp:
            tmp.write(data)
            tmp_path = tmp.name

        try:
            result = subprocess.run(
                [
                    "pg_restore",
                    "-h", db_info["host"],
                    "-p", db_info["port"],
                    "-U", db_info["user"],
                    "-d", db_info["dbname"],
                    "--no-password",
                    "--clean",
                    "--if-exists",
                    tmp_path,
                ],
                env=env,
                capture_output=True,
                timeout=120,
            )
            stderr = result.stderr.decode()
            if result.returncode != 0 and "error" in stderr.lower():
                logger.error("pg_restore failed: %s", stderr)
                raise HTTPException(
                    502,
                    detail={"code": "PG_RESTORE_ERROR", "message": stderr},
                )
            logger.info("PostgreSQL DB restored by user %s", current_user.id)
            return {"ok": True, "message": "База данных успешно восстановлена."}
        except FileNotFoundError:
            raise HTTPException(
                500,
                detail={"code": "PG_RESTORE_MISSING", "message": "pg_restore не установлен"},
            )
        finally:
            os.unlink(tmp_path)


# ─── Calendar Excel Export ────────────────────────────────────────────────────

@router.get("/calendar/export")
@limiter.limit("20/minute")
async def export_calendar_excel(
    request: Request,
    from_date: date = Query(None, alias="from"),
    to_date: date = Query(None, alias="to"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Export bookings calendar to Excel (.xlsx). Requires admin role."""
    logger.info("Calendar export requested by user %s", current_user.id)

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            500,
            detail={"code": "OPENPYXL_MISSING", "message": "openpyxl не установлен на сервере"},
        )

    # Fetch bookings
    q = select(Booking).order_by(Booking.start_date)
    if from_date:
        q = q.where(Booking.end_date >= from_date)
    if to_date:
        q = q.where(Booking.start_date <= to_date)
    result = await db.execute(q)
    bookings = result.scalars().all()

    # Fetch rooms for lookup
    rooms_res = await db.execute(select(Room).order_by(Room.number))
    rooms = {r.id: r for r in rooms_res.scalars().all()}

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Бронирования"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill("solid", fgColor="F0F4F8")

    columns = [
        ("№ Номера", 12),
        ("Этаж", 8),
        ("Гость", 22),
        ("Телефон", 16),
        ("Дата заезда", 14),
        ("Дата выезда", 14),
        ("Ночей", 8),
        ("Примечания", 30),
        ("Создано", 18),
    ]

    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

    for row_idx, booking in enumerate(bookings, start=2):
        room = rooms.get(booking.room_id)
        nights = (booking.end_date - booking.start_date).days
        row_data = [
            room.number if room else "—",
            room.floor if room else "—",
            booking.guest_name,
            booking.phone or "",
            booking.start_date.strftime("%d.%m.%Y"),
            booking.end_date.strftime("%d.%m.%Y"),
            nights,
            booking.notes or "",
            booking.created_at.strftime("%d.%m.%Y %H:%M") if booking.created_at else "",
        ]
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if fill:
                cell.fill = fill
            cell.alignment = center_align if col_idx in (1, 2, 5, 6, 7) else cell_align
        ws.row_dimensions[row_idx].height = 20

    summary_row = len(bookings) + 2
    ws.cell(row=summary_row, column=1, value=f"Всего бронирований: {len(bookings)}")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, italic=True, color="555555")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = "calendar"
    if from_date:
        filename += f"_from_{from_date}"
    if to_date:
        filename += f"_to_{to_date}"
    filename += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
